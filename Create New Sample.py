import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from datetime import datetime, date



def create_excel_file(data):
    current_year = datetime.now().year
    current_month = datetime.now().strftime("%B")
    file_name = f"Lab Sample {current_year}.xlsx"

    try:
        wb = load_workbook(file_name)
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.title = current_month
    else:
        if current_month in wb.sheetnames:
            sheet = wb[current_month]
        else:
            sheet = wb.create_sheet(title=current_month)

    if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
        # Write headers if the sheet is empty
        headers = [
            "No", "User's Name", "Sample ID", "Customer Name", "Customer Contact Number",
            "Sample Name", "Sample Type", "Sample Condition", "Test Request", "Test Type",
            "Date Received", "Remark 1", "Date Commence", "Date Completion", "Date Issued", "Remark 2"
        ]
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num).value = header

    # Write data
    next_row = sheet.max_row + 1
    headers = [
        "No", "User's Name", "Sample ID", "Customer Name", "Customer Contact Number",
        "Sample Name", "Sample Type", "Sample Condition", "Test Request", "Test Type",
        "Date Received", "Remark 1", "Date Commence", "Date Completion", "Date Issued", "Remark 2"
    ]
    # Modify the block that writes the data to include the Sample ID field
    for col_num, header in enumerate(headers, start=1):
        if header == "No":
            sheet.cell(row=next_row, column=col_num).value = next_row - 1
        elif header == "User's Name":
            sheet.cell(row=next_row, column=col_num).value = data.get('user_name', '')
        elif header == "Sample ID":
            column_a_value = sheet.cell(row=next_row, column=1).value
            sample_id = f"{data.get('sample_id', '')}{column_a_value}"
            sheet.cell(row=next_row, column=col_num).value = sample_id
        elif header == "Customer Name":
            sheet.cell(row=next_row, column=col_num).value = data.get('customer_name', '')
        elif header == "Customer Contact Number":
            sheet.cell(row=next_row, column=col_num).value = data.get('customer_contact', '')
        elif header == "Sample Name":
            sheet.cell(row=next_row, column=col_num).value = data.get('sample_name', '')
        elif header == "Sample Type":
            sheet.cell(row=next_row, column=col_num).value = data.get('sample_type', '')
        elif header == "Sample Condition":
            sheet.cell(row=next_row, column=col_num).value = data.get('sample_condition', '')
        elif header == "Test Request":
            sheet.cell(row=next_row, column=col_num).value = ', '.join(data.get('test_requested', []))
        elif header == "Test Type":
            sheet.cell(row=next_row, column=col_num).value = data.get('type_of_test', '')
        elif header == "Date Received":
            sheet.cell(row=next_row, column=col_num).value = datetime.now().strftime("%Y-%m-%d")
        elif header == "Remark 1":
            sheet.cell(row=next_row, column=col_num).value = data.get('remark', '')
        elif header == "Date Commence":
            sheet.cell(row=next_row, column=col_num).value = data.get('date_commence', '')
        elif header == "Date Completion":
            sheet.cell(row=next_row, column=col_num).value = data.get('date_completion', '')
        elif header == "Date Issued":
            sheet.cell(row=next_row, column=col_num).value = data.get('date_issued', '')
        elif header == "Remark 2":
            sheet.cell(row=next_row, column=col_num).value = data.get('remark_2', '')

    wb.save(file_name)
    messagebox.showinfo("Success", f"Excel file saved at: {file_name}")


sample_condition_var = None
test_type_var = None


def submit_form():
    # Generate Sample ID
    current_date = date.today().strftime("%d%m%y")
    type_of_test = test_type_var.get()
    if type_of_test == "MAL":
        sample_id = f"{current_date}M"
    elif type_of_test == "FOOD":
        sample_id = f"{current_date}F"

    # Retrieve form data
    data = {
        'user_name': name_entry.get(),
        'sample_name': sample_name_entry.get(),
        'sample_type': sample_type_entry.get(),
        'customer_name': customer_name_entry.get(),
        'customer_contact': customer_contact_entry.get(),
        'sample_condition': sample_condition_var.get(),  # Retrieve the selected sample condition
        'test_requested': [],
        'type_of_test': test_type_var.get(),  # Retrieve the selected type of test
        'remark': remark_entry.get(),
        'sample_id': sample_id
    }

        
    # Retrieve selected test requests
    for test, var in test_requested_vars.items():
        if var.get():
            data['test_requested'].append(test)

    create_excel_file(data)

def main():
    global name_entry, sample_name_entry, sample_type_entry, customer_name_entry, customer_contact_entry, remark_entry
    global sample_condition_var, test_requested_vars, test_type_var

    # Create GUI window
    window = tk.Tk()
    window.title("New Sample Regristration")

    # Create form labels and entry fields
    tk.Label(window, text="User's Name:").grid(row=0, column=0, sticky='w')
    name_entry = tk.Entry(window)
    name_entry.grid(row=0, column=1)

    tk.Label(window, text="Sample Name:").grid(row=1, column=0, sticky='w')
    sample_name_entry = tk.Entry(window)
    sample_name_entry.grid(row=1, column=1)

    tk.Label(window, text="Sample Type:").grid(row=2, column=0, sticky='w')
    sample_type_entry = tk.Entry(window)
    sample_type_entry.grid(row=2, column=1)

    tk.Label(window, text="Customer Name:").grid(row=3, column=0, sticky='w')
    customer_name_entry = tk.Entry(window)
    customer_name_entry.grid(row=3, column=1)

    tk.Label(window, text="Customer Contact Number:").grid(row=4, column=0, sticky='w')
    customer_contact_entry = tk.Entry(window)
    customer_contact_entry.grid(row=4, column=1)

    # Create sample condition radio buttons
    tk.Label(window, text="Sample Condition:").grid(row=5, column=0, sticky='w')
    sample_conditions = ["Acceptable", "Broken", "Other"]
    sample_condition_var = tk.StringVar(value="")  # Initialize with an empty value
    for i, condition in enumerate(sample_conditions, start=0):
        radio_button = tk.Radiobutton(window, text=condition, variable=sample_condition_var, value=condition)
        radio_button.grid(row=i + 5, column=1, sticky='w')

    # Create test requested checkboxes
    test_requested_vars = {}
    tk.Label(window, text="Test Requested:").grid(row=8, column=0, sticky='w')
    test_requested = [
        "Aerobic Plate Count", "Yeast and Moulds", "Bile-tolerant gram-negative bacteria",
        "Escherichia coli", "Salmonella species", "Staphylococcus aureus",
        "Bifidobacterium", "Lactobacillus", "Streptococcus thermophilus"
    ]
    for i, test in enumerate(test_requested, start=0):
        var = tk.IntVar()
        checkbox = tk.Checkbutton(window, text=test, variable=var)
        checkbox.grid(row=i+8, column=1, sticky='w')
        test_requested_vars[test] = var

    # Create type of test radio buttons
    test_type_var = tk.StringVar(value="")  # Initialize with an empty value
    tk.Label(window, text="Type of Test:").grid(row=17, column=0, sticky='w')
    test_types = ["MAL", "FOOD"]
    for i, test_type in enumerate(test_types, start=0):
        radio_button = tk.Radiobutton(window, text=test_type, variable=test_type_var, value=test_type)
        radio_button.grid(row=i + 17, column=1, sticky='w')

    tk.Label(window, text="Remark:").grid(row=19, column=0, sticky='w')
    remark_entry = tk.Entry(window)
    remark_entry.grid(row=19, column=1)

    submit_button = tk.Button(window, text="Submit", command=submit_form)
    submit_button.grid(row=20, column=1, pady=10)

    # Start the main event loop
    window.mainloop()


if __name__ == "__main__":
    main()

