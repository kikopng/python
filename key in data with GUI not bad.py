import openpyxl
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import filedialog

def search_sample():
    sample_id = sample_id_entry.get()
    if sample_id:
        file_path = file_path_entry.get()
        if file_path:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Get the dimensions of the used range
            rows_count = sheet.max_row
            cols_count = sheet.max_column

            # Iterate over each cell in the used range and search for the sample ID
            found = False
            item = ""
            for row in range(1, rows_count + 1):
                for col in range(1, cols_count + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None and str(sample_id) in str(cell.value):
                        found = True
                        item = sheet.cell(row=row, column=9).value  # Get the item from the 9th column
                        break
                if found:
                    break

            if found:
                bacterial_species = [species.strip() for species in item.split(',')]

                # Create a new workbook for storing the results
                new_file_name = os.path.splitext(file_path)[0] + "_data.xlsx"
                new_workbook = load_workbook(new_file_name) if os.path.exists(new_file_name) else openpyxl.Workbook()
                new_sheet = new_workbook.active
                headers = ["Sample ID", "Species", "Result"]
                header_row = headers + ["Date Commence:", "", "Date Completion:"]
                new_sheet.append(header_row)
                

                for species in bacterial_species:
                    if species == "Bile-tolerant gram-negative bacteria":
                        ask_bile_tolerant_presence(new_sheet, sample_id)
                    elif species in ["Escherichia coli", "Salmonella species", "Staphylococcus aureus"]:
                        ask_bacterial_presence(new_sheet, sample_id, species)
                    elif species in ["Aerobic Plate Count", "Yeast and Moulds", "Bifidobacterium", "Lactobacillus", "Streptococcus thermophilus"]:
                        ask_dilution_colony_count(new_sheet, sample_id, species)

                new_workbook.save(new_file_name)  # Save the new workbook
                result_text.set("Results have been saved to the new Excel file: " + new_file_name)
            else:
                result_text.set("Sample ID not found.")
        else:
            result_text.set("No file selected.")
    else:
        result_text.set("Please enter a sample ID.")


from tkinter.simpledialog import Dialog
from tkinter import Label, IntVar, Radiobutton, Button

class BileTolerantDialog(Dialog):
    def __init__(self, parent, title=None):
        super().__init__(parent, title)
        self.presence_flags = []

    def body(self, master):
        weights = ["0.1 g", "0.01 g", "0.001 g"]
        Label(master, text="Bile-tolerant bacteria:").grid(row=0, column=0, columnspan=3)

        self.vars = []
        for i, weight in enumerate(weights):
            var = IntVar()
            Label(master, text=weight + ":").grid(row=i+1, column=0, sticky="e")
            Radiobutton(master, variable=var, value=0, text="Absence").grid(row=i+1, column=1)
            Radiobutton(master, variable=var, value=1, text="Presence").grid(row=i+1, column=2)
            self.vars.append(var)

    def validate(self):
        self.presence_flags = [var.get() == 1 for var in self.vars]
        
        # Check for valid combinations
        if self.presence_flags == [False, False, False]:
            # Absence for all three weights
            return True
        elif self.presence_flags == [True, True, True]:
            # Presence for all three weights
            return True
        elif self.presence_flags == [True, True, False]:
            # Presence for the first two weights and absence for the last weight
            return True
        elif self.presence_flags == [True, False, False]:
            # Presence for the first weight and absence for the last two weights
            return True
        else:
            return False

    def apply(self):
        pass

def ask_bile_tolerant_presence(sheet, sample_id):
    result_text.set("Bile-tolerant gram-negative bacteria:")
    dialog = BileTolerantDialog(window, "Bile-tolerant gram-negative bacteria")

    while not dialog.validate():
        if not dialog.ok:
            return
        else:
            result_text.set("Please select at least one presence.")

        dialog = BileTolerantDialog(window, "Bile-tolerant gram-negative bacteria")

    presence_flags = dialog.presence_flags


    if all(presence_flags):
        result = ">1000"
    elif presence_flags[0] and presence_flags[1] and not presence_flags[2]:
        result = "<1000 and >100"
    elif presence_flags[0] and not presence_flags[1] and not presence_flags[2]:
        result = "<100 and >10"
    elif not presence_flags[0] and not presence_flags[1] and not presence_flags[2]:
        result = "<10"

    row = [sample_id, "Bile-tolerant gram-negative bacteria", result]
    sheet.append(row)


def ask_bacterial_presence(sheet, sample_id, species):
    presence = tk.messagebox.askyesno(species + ": Presence", "Is " + species + " present?")

    if presence:
        presence_str = "Presence"
    else:
        presence_str = "Absence"

    # Write the result in the new sheet
    row = [sample_id, species, presence_str]
    sheet.append(row)


from tkinter.simpledialog import Dialog
from tkinter import Label, Entry, Button

class CustomInputDialog(Dialog):
    def __init__(self, parent, title=None, species=None):
        self.species = species  # Store the species information
        super().__init__(parent, title)

    def body(self, master):
        species_label = Label(master, text="Bacterial Species:")
        species_label.grid(row=0, column=0, sticky="w")
        species_value_label = Label(master, text=self.species)
        species_value_label.grid(row=0, column=1, sticky="w")

        Label(master, text="Enter the dilutions and colony numbers:").grid(row=1, column=0, columnspan=2)
        Label(master, text="Dilution 1:").grid(row=2, column=0)
        Label(master, text="Colony Number 1:").grid(row=3, column=0)
        Label(master, text="Colony Number 2:").grid(row=4, column=0)
        Label(master, text="Dilution 2:").grid(row=2, column=2)
        Label(master, text="Colony Number 1:").grid(row=3, column=2)
        Label(master, text="Colony Number 2:").grid(row=4, column=2)

        self.dilution1_entry = Entry(master)
        self.dilution1_entry.grid(row=2, column=1)
        self.colony1_entry = Entry(master)
        self.colony1_entry.grid(row=3, column=1)
        self.colony2_entry = Entry(master)
        self.colony2_entry.grid(row=4, column=1)
        self.dilution2_entry = Entry(master)
        self.dilution2_entry.grid(row=2, column=3)
        self.colony3_entry = Entry(master)
        self.colony3_entry.grid(row=3, column=3)
        self.colony4_entry = Entry(master)
        self.colony4_entry.grid(row=4, column=3)

    def validate(self):
        self.dilution1 = self.validate_input(self.dilution1_entry.get())
        self.colony1 = self.validate_input(self.colony1_entry.get())
        self.colony2 = self.validate_input(self.colony2_entry.get())
        self.dilution2 = self.validate_input(self.dilution2_entry.get())
        self.colony3 = self.validate_input(self.colony3_entry.get())
        self.colony4 = self.validate_input(self.colony4_entry.get())

        if None in (self.dilution1, self.colony1, self.colony2, self.dilution2, self.colony3, self.colony4):
            return False
        return True

    def validate_input(self, value):
        if value.strip() == "":
            return 0
        try:
            return float(value)
        except ValueError:
            return None


def ask_dilution_colony_count(sheet, sample_id, species):
    result_text.set(species + ":")
    
    input_dialog = CustomInputDialog(window, "Enter Dilutions and Colony Numbers", species)


    if input_dialog.ok:
        dilution1 = input_dialog.dilution1
        dilution2 = input_dialog.dilution2

        # Handle colony numbers and count
        colony_numbers1 = [input_dialog.colony1, input_dialog.colony2]
        colony_numbers2 = [input_dialog.colony3, input_dialog.colony4]

        # Exclude blank colony numbers from the count
        no_cn_in_fd = sum(1 for colony in colony_numbers1 if colony != 0)
        no_cn_in_sd = sum(1 for colony in colony_numbers2 if colony != 0)

        # Determine larger and smaller dilutions
        if dilution1 > dilution2:
            first_dilution = dilution1
            second_dilution = dilution2
        else:
            first_dilution = dilution2
            second_dilution = dilution1

        # Calculate bacterial count
        colony_sum = sum(map(float, colony_numbers1)) + sum(map(float, colony_numbers2))
        bacterial_count = int(round(colony_sum / ((no_cn_in_fd + 0.1 * no_cn_in_sd) * first_dilution), -1))

        # Write the result in the new sheet
        row = [sample_id, species, bacterial_count]
        sheet.append(row)



# Create the main application window
window = tk.Tk()
window.title("Sample Search")

# Create and place the widgets
sample_id_label = tk.Label(window, text="Sample ID:")
sample_id_label.pack()

sample_id_entry = tk.Entry(window)
sample_id_entry.pack()

file_path_label = tk.Label(window, text="Excel file path:")
file_path_label.pack()

file_path_entry = tk.Entry(window)
file_path_entry.pack()

select_file_button = tk.Button(window, text="Select File", command=lambda: file_path_entry.insert(tk.END, filedialog.askopenfilename()))
select_file_button.pack()

search_button = tk.Button(window, text="Search", command=search_sample)
search_button.pack()

result_text = tk.StringVar()
result_label = tk.Label(window, textvariable=result_text)
result_label.pack()

# Run the Tkinter event loop
window.mainloop()
